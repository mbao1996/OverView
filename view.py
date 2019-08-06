# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from viewlib import *

# code cab be assigend
code = '600036'

try:
    wb = load_workbook(fn, keep_vba=True)
except Exception as e:
    print(str(e))
    os._exit(0)
ws = wb['Sheet1']
rd = RawData()
rd.reset(rd)

# get name, price
ws.cell(row_viewname, col_code).value = get_sina_id(code)
name, price = get_name_price(code)
ws.cell(row_viewname, col_share_name).value = name
ws.cell(row_viewname, col_price).value = price

year_title(ws, rd, code, row_year, col_start)
# ROE,企业自由现金流量
query(ws, rd, code)
# 期末总股本
balancesheet(ws, rd, code)
# 营业总收入,营业利润,净利润
income(ws, rd, code)
grow(ws, row_total_revenue, u'营总收增长率')
grow(ws, row_net_income, u'净利润增长率')

try:
    wb.save(fn)
except Exception as e:
    print(str(e))
    os._exit(0)
    
print('\n finished')
